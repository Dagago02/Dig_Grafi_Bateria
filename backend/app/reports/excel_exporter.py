import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from app.models.evaluation import Evaluation
from app.models.participant import Participant
from app.models.result import Result
from app.calculations.calculator import calculate_participant_results

def generate_evaluation_excel(evaluation_id: int, db: Session) -> io.BytesIO:
    """
    Genera un archivo Excel (.xlsx) con la estructura exacta de Ejemplo.xlsx.
    """
    evaluation = db.query(Evaluation).filter(Evaluation.id == evaluation_id).first()
    if not evaluation:
        raise ValueError(f"Evaluación id {evaluation_id} no encontrada.")

    participants = db.query(Participant).filter(Participant.evaluacion_id == evaluation_id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados Batería"

    headers = [
        "Fecha de aplicación", "Año de aplicación", "Número de identificación", "Nombre completo", "Sexo",
        "Año de nacimiento", "Estado civil", "Último nivel de estudios", "Ocupación / profesión",
        "Lugar residencia - Ciudad", "Lugar residencia - Depto", "Estrato", "Tipo de vivienda",
        "Personas a cargo", "Lugar trabajo - Ciudad", "Lugar trabajo - Depto", "Años en la empresa",
        "Nombre del cargo", "Tipo de cargo (nivel)", "Años en el cargo", "Área / Sección",
        "Tipo de contrato", "Horas diarias", "Tipo de salario",

        # Forma A
        "Dimensión: Características del liderazgo - Forma A (puntaje transformado)",
        "Dimensión: Características del liderazgo - Forma A (nivel de riesgo)",
        "Dimensión: Relaciones sociales en el trabajo - Forma A (puntaje transformado)",
        "Dimensión: Relaciones sociales en el trabajo - Forma A (nivel de riesgo)",
        "Dimensión: Retroalimentación del desempeño - Forma A (puntaje transformado)",
        "Dimensión: Retroalimentación del desempeño - Forma A (nivel de riesgo)",
        "Dimensión: Relación con los colaboradores - Forma A (puntaje transformado)",
        "Dimensión: Relación con los colaboradores - Forma A (nivel de riesgo)",
        "DOMINIO: Liderazgo y relaciones sociales en el trabajo - Forma A (puntaje transformado)",
        "DOMINIO: Liderazgo y relaciones sociales en el trabajo - Forma A (nivel de riesgo)",
        "Dimensión: Claridad de rol - Forma A (puntaje transformado)",
        "Dimensión: Claridad de rol - Forma A (nivel de riesgo)",
        "Dimensión: Capacitación - Forma A (puntaje transformado)",
        "Dimensión: Capacitación - Forma A (nivel de riesgo)",
        "Dimensión: Participación y manejo del cambio - Forma A (puntaje transformado)",
        "Dimensión: Participación y manejo del cambio - Forma A (nivel de riesgo)",
        "Dimensión: Oportunidades para el uso y desarrollo de habilidades y conocimientos - Forma A (puntaje transformado)",
        "Dimensión: Oportunidades para el uso y desarrollo de habilidades y conocimientos - Forma A (nivel de riesgo)",
        "Dimensión: Control y autonomía sobre el trabajo - Forma A (puntaje transformado)",
        "Dimensión: Control y autonomía sobre el trabajo - Forma A (nivel de riesgo)",
        "DOMINIO: Control sobre el trabajo - Forma A (puntaje transformado)",
        "DOMINIO: Control sobre el trabajo - Forma A (nivel de riesgo)",
        "Dimensión: Demandas ambientales y de esfuerzo físico - Forma A (puntaje transformado)",
        "Dimensión: Demandas ambientales y de esfuerzo físico - Forma A (nivel de riesgo)",
        "Dimensión: Demandas emocionales - Forma A (puntaje transformado)",
        "Dimensión: Demandas emocionales - Forma A (nivel de riesgo)",
        "Dimensión: Demandas cuantitativas - Forma A (puntaje transformado)",
        "Dimensión: Demandas cuantitativas - Forma A (nivel de riesgo)",
        "Dimensión: Influencia del trabajo sobre el entorno extralaboral - Forma A (puntaje transformado)",
        "Dimensión: Influencia del trabajo sobre el entorno extralaboral - Forma A (nivel de riesgo)",
        "Dimensión: Exigencias de responsabilidad del cargo - Forma A (puntaje transformado)",
        "Dimensión: Exigencias de responsabilidad del cargo - Forma A (nivel de riesgo)",
        "Dimensión: Demandas de carga mental - Forma A (puntaje transformado)",
        "Dimensión: Demandas de carga mental - Forma A (nivel de riesgo)",
        "Dimensión: Consistencia del rol - Forma A (puntaje transformado)",
        "Dimensión: Consistencia del rol - Forma A (nivel de riesgo)",
        "Dimensión: Demandas de la jornada de trabajo - Forma A (puntaje transformado)",
        "Dimensión: Demandas de la jornada de trabajo - Forma A (nivel de riesgo)",
        "DOMINIO: Demandas del trabajo - Forma A (puntaje transformado)",
        "DOMINIO: Demandas del trabajo - Forma A (nivel de riesgo)",
        "Dimensión: Recompensas derivadas de la pertenencia a la organización y del trabajo que se realiza - Forma A (puntaje transformado)",
        "Dimensión: Recompensas derivadas de la pertenencia a la organización y del trabajo que se realiza - Forma A (nivel de riesgo)",
        "Dimensión: Reconocimiento y compensación - Forma A (puntaje transformado)",
        "Dimensión: Reconocimiento y compensación - Forma A (nivel de riesgo)",
        "DOMINIO: Recompensas - Forma A (puntaje transformado)",
        "DOMINIO: Recompensas - Forma A (nivel de riesgo)",
        "PUNTAJE TOTAL Intralaboral - Forma A (puntaje transformado)",
        "PUNTAJE TOTAL Intralaboral - Forma A (nivel de riesgo)",

        # Forma B
        "Dimensión: Características del liderazgo - Forma B (puntaje transformado)",
        "Dimensión: Características del liderazgo - Forma B (nivel de riesgo)",
        "Dimensión: Relaciones sociales en el trabajo - Forma B (puntaje transformado)",
        "Dimensión: Relaciones sociales en el trabajo - Forma B (nivel de riesgo)",
        "Dimensión: Retroalimentación del desempeño - Forma B (puntaje transformado)",
        "Dimensión: Retroalimentación del desempeño - Forma B (nivel de riesgo)",
        "DOMINIO: Liderazgo y relaciones sociales en el trabajo - Forma B (puntaje transformado)",
        "DOMINIO: Liderazgo y relaciones sociales en el trabajo - Forma B (nivel de riesgo)",
        "Dimensión: Claridad de rol - Forma B (puntaje transformado)",
        "Dimensión: Claridad de rol - Forma B (nivel de riesgo)",
        "Dimensión: Capacitación - Forma B (puntaje transformado)",
        "Dimensión: Capacitación - Forma B (nivel de riesgo)",
        "Dimensión: Participación y manejo del cambio - Forma B (puntaje transformado)",
        "Dimensión: Participación y manejo del cambio - Forma B (nivel de riesgo)",
        "Dimensión: Oportunidades para el uso y desarrollo de habilidades y conocimientos - Forma B (puntaje transformado)",
        "Dimensión: Oportunidades para el uso y desarrollo de habilidades y conocimientos - Forma B (nivel de riesgo)",
        "Dimensión: Control y autonomía sobre el trabajo - Forma B (puntaje transformado)",
        "Dimensión: Control y autonomía sobre el trabajo - Forma B (nivel de riesgo)",
        "DOMINIO: Control sobre el trabajo - Forma B (puntaje transformado)",
        "DOMINIO: Control sobre el trabajo - Forma B (nivel de riesgo)",
        "Dimensión: Demandas ambientales y de esfuerzo físico - Forma B (puntaje transformado)",
        "Dimensión: Demandas ambientales y de esfuerzo físico - Forma B (nivel de riesgo)",
        "Dimensión: Demandas emocionales - Forma B (puntaje transformado)",
        "Dimensión: Demandas emocionales - Forma B (nivel de riesgo)",
        "Dimensión: Demandas cuantitativas - Forma B (puntaje transformado)",
        "Dimensión: Demandas cuantitativas - Forma B (nivel de riesgo)",
        "Dimensión: Influencia del trabajo sobre el entorno extralaboral - Forma B (puntaje transformado)",
        "Dimensión: Influencia del entorno extralaboral - Forma B (nivel de riesgo)",
        "Dimensión: Demandas de carga mental - Forma B (puntaje transformado)",
        "Dimensión: Demandas de carga mental - Forma B (nivel de riesgo)",
        "Dimensión: Demandas de la jornada de trabajo - Forma B (puntaje transformado)",
        "Dimensión: Demandas de la jornada de trabajo - Forma B (nivel de riesgo)",
        "DOMINIO: Demandas del trabajo - Forma B (puntaje transformado)",
        "DOMINIO: Demandas del trabajo - Forma B (nivel de riesgo)",
        "Dimensión: Recompensas derivadas de la pertenencia a la organización y del trabajo que se realiza - Forma B (puntaje transformado)",
        "Dimensión: Recompensas derivadas de la pertenencia a la organización y del trabajo que se realiza - Forma B (nivel de riesgo)",
        "Dimensión: Reconocimiento y compensación - Forma B (puntaje transformado)",
        "Dimensión: Reconocimiento y compensación - Forma B (nivel de riesgo)",
        "DOMINIO: Recompensas - Forma B (puntaje transformado)",
        "DOMINIO: Recompensas - Forma B (nivel de riesgo)",
        "PUNTAJE TOTAL Intralaboral - Forma B (puntaje transformado)",
        "PUNTAJE TOTAL Intralaboral - Forma B (nivel de riesgo)",

        # Extralaboral
        "Dimensión: Tiempo fuera del trabajo - Extralaboral (puntaje transformado)",
        "Dimensión: Tiempo fuera del trabajo - Extralaboral (nivel de riesgo)",
        "Dimensión: Relaciones familiares - Extralaboral (puntaje transformado)",
        "Dimensión: Relaciones familiares - Extralaboral (nivel de riesgo)",
        "Dimensión: Comunicación y relaciones interpersonales - Extralaboral (puntaje transformado)",
        "Dimensión: Comunicación y relaciones interpersonales - Extralaboral (nivel de riesgo)",
        "Dimensión: Situación económica del grupo familiar - Extralaboral (puntaje transformado)",
        "Dimensión: Situación económica del grupo familiar - Extralaboral (nivel de riesgo)",
        "Dimensión: Características de la vivienda y de su entorno - Extralaboral (puntaje transformado)",
        "Dimensión: Características de la vivienda y de su entorno - Extralaboral (nivel de riesgo)",
        "Dimensión: Influencia del entorno extralaboral sobre el trabajo - Extralaboral (puntaje transformado)",
        "Dimensión: Influencia del entorno extralaboral sobre el trabajo - Extralaboral (nivel de riesgo)",
        "Dimensión: Desplazamiento vivienda - trabajo - vivienda - Extralaboral (puntaje transformado)",
        "Dimensión: Desplazamiento vivienda - trabajo - vivienda - Extralaboral (nivel de riesgo)",
        "PUNTAJE TOTAL Extralaboral (puntaje transformado)",
        "PUNTAJE TOTAL Extralaboral (nivel de riesgo)",

        # Estrés & Total General
        "PUNTAJE TOTAL Estrés (puntaje transformado)",
        "PUNTAJE TOTAL Estrés (nivel de riesgo)",
        "PUNTAJE TOTAL GENERAL (puntaje transformado)",
        "PUNTAJE TOTAL GENERAL (nivel de riesgo)"
    ]

    ws.append(headers)

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for part in participants:
        try:
            calculate_participant_results(part.id, db)
        except Exception as e:
            print(f"Error procesando resultados para {part.id}: {e}")

        results_list = db.query(Result).filter(Result.participant_id == part.id).all()
        res_map = {(r.tipo_resultado, r.nombre_target): r for r in results_list}

        is_forma_a = (part.tipo_forma == "A")

        row = [
            evaluation.created_at.strftime("%d/%m/%Y") if evaluation.created_at else "",
            evaluation.fecha_inicio.year if evaluation.fecha_inicio else 2026,
            part.cedula,
            f"{part.nombres} {part.apellidos}",
            part.sexo or "",
            (2026 - part.edad) if part.edad else "",
            part.estado_civil or "",
            part.nivel_educativo or "",
            part.cargo or "",
            "Bogotá", "Cundinamarca",
            3, "Propia", 1, "Bogotá", "Cundinamarca",
            part.tiempo_empresa or "",
            part.cargo or "",
            "Profesional, analista, técnico especializado" if is_forma_a else "Auxiliar, asistente administrativo",
            part.tiempo_empresa or "",
            part.area or "",
            part.tipo_contrato or "Término indefinido",
            8, "Fijo"
        ]

        def get_sr(tipo: str, nom: str):
            r = res_map.get((tipo, nom))
            if r:
                return r.puntaje_transformado, r.nivel_riesgo
            return None, None

        # Forma A Items
        forma_a_items = [
            ("dimension", "Características del liderazgo"),
            ("dimension", "Relaciones sociales en el trabajo"),
            ("dimension", "Retroalimentación del desempeño"),
            ("dimension", "Relación con los colaboradores"),
            ("dominio", "Liderazgo y relaciones sociales en el trabajo"),
            ("dimension", "Claridad de rol"),
            ("dimension", "Capacitación"),
            ("dimension", "Participación y manejo del cambio"),
            ("dimension", "Oportunidades para el uso y desarrollo de habilidades y conocimientos"),
            ("dimension", "Control y autonomía sobre el trabajo"),
            ("dominio", "Control sobre el trabajo"),
            ("dimension", "Demandas ambientales y de esfuerzo físico"),
            ("dimension", "Demandas emocionales"),
            ("dimension", "Demandas cuantitativas"),
            ("dimension", "Influencia del trabajo sobre el entorno extralaboral"),
            ("dimension", "Exigencias de responsabilidad del cargo"),
            ("dimension", "Demandas de carga mental"),
            ("dimension", "Consistencia del rol"),
            ("dimension", "Demandas de la jornada de trabajo"),
            ("dominio", "Demandas del trabajo"),
            ("dimension", "Recompensas derivadas de la pertenencia a la organización"),
            ("dimension", "Reconocimiento y compensación"),
            ("dominio", "Recompensas"),
            ("general", "intralaboral"),
        ]

        for tipo, nom in forma_a_items:
            if is_forma_a:
                sc, rsk = get_sr(tipo, nom)
                row.extend([sc, rsk])
            else:
                row.extend([None, None])

        # Forma B Items
        forma_b_items = [
            ("dimension", "Características del liderazgo"),
            ("dimension", "Relaciones sociales en el trabajo"),
            ("dimension", "Retroalimentación del desempeño"),
            ("dominio", "Liderazgo y relaciones sociales en el trabajo"),
            ("dimension", "Claridad de rol"),
            ("dimension", "Capacitación"),
            ("dimension", "Participación y manejo del cambio"),
            ("dimension", "Oportunidades para el uso y desarrollo de habilidades y conocimientos"),
            ("dimension", "Control y autonomía sobre el trabajo"),
            ("dominio", "Control sobre el trabajo"),
            ("dimension", "Demandas ambientales y de esfuerzo físico"),
            ("dimension", "Demandas emocionales"),
            ("dimension", "Demandas cuantitativas"),
            ("dimension", "Influencia del trabajo sobre el entorno extralaboral"),
            ("dimension", "Demandas de carga mental"),
            ("dimension", "Demandas de la jornada de trabajo"),
            ("dominio", "Demandas del trabajo"),
            ("dimension", "Recompensas derivadas de la pertenencia a la organización"),
            ("dimension", "Reconocimiento y compensación"),
            ("dominio", "Recompensas"),
            ("general", "intralaboral"),
        ]

        for tipo, nom in forma_b_items:
            if not is_forma_a:
                sc, rsk = get_sr(tipo, nom)
                row.extend([sc, rsk])
            else:
                row.extend([None, None])

        # Extralaboral Items
        extra_items = [
            ("dimension", "Tiempo fuera del trabajo"),
            ("dimension", "Relaciones familiares"),
            ("dimension", "Comunicación y relaciones interpersonales"),
            ("dimension", "Situación económica del grupo familiar"),
            ("dimension", "Características de la vivienda y de su entorno"),
            ("dimension", "Influencia del entorno extralaboral sobre el trabajo"),
            ("dimension", "Desplazamiento vivienda-trabajo-vivienda"),
            ("general", "extralaboral")
        ]

        for tipo, nom in extra_items:
            sc, rsk = get_sr(tipo, nom)
            row.extend([sc, rsk])

        # Estrés & Total General
        sc_estres, rsk_estres = get_sr("general", "estres")
        row.extend([sc_estres, rsk_estres])

        sc_gen, rsk_gen = get_sr("general", "total_general")
        row.extend([sc_gen, rsk_gen])

        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
